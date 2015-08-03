
from openpyxl import load_workbook
from openpyxl.charts import Reference, Series, PieChart
from openpyxl.styles.colors import GREEN

import requests
import datetime

MONITOR_ID_COLUMN = 'K'
MONITOR_ID_DATA_STARTS = 2

DATA_COLUMN = 'G'
FIRST_DATA_TABLE_ROW = 4
TOTAL_DOWNTIME_ROW_OFFSET = 3
NEXT_TABLE_ROW_OFFSET = 13


def main():
    #Loading file with configurations
    #Each sheet represents a server and it has the id of all the server monitors
    wb = load_workbook('../templates/newtemplate.xlsx')
    print "Servers to Process:"
    print wb.get_sheet_names()


    for server_name in wb.sheetnames:
        monitor_first_row = FIRST_DATA_TABLE_ROW
        sheet = wb[server_name]
        monitor_id_line = MONITOR_ID_DATA_STARTS
        monitor_id = sheet[MONITOR_ID_COLUMN+str(monitor_id_line)].value
        i = 0;
        while monitor_id:
            url = 'http://172.18.42.167/AppManager/json/GetDowntimeDetails?apikey=e7809b7a69757632d773572dd049aeac&period=11&resourceid=%s&showFullDetails=false' % str(monitor_id)
            print url
            r = requests.get(url)
            print r.json()
            try:
                down_percent = r.json()['response']['result'][0]['DownPercent']
                sheet[DATA_COLUMN+str(monitor_first_row + TOTAL_DOWNTIME_ROW_OFFSET)] = float(down_percent)/100
                add_pie_chart(sheet,monitor_first_row,i)
            except Exception as e:
                print e

            monitor_id_line += 1
            i += 1
            monitor_id = sheet[MONITOR_ID_COLUMN+str(monitor_id_line)].value
            monitor_first_row += NEXT_TABLE_ROW_OFFSET


    wb.save('../output/AppManager_DowntimeSummary_Monthly.xlsx' )



def add_pie_chart(sheet, write_line,i):
    yvalues = Reference(sheet, (write_line+TOTAL_DOWNTIME_ROW_OFFSET,7),(write_line+TOTAL_DOWNTIME_ROW_OFFSET+1,7))
    xvalues = Reference(sheet, (write_line+TOTAL_DOWNTIME_ROW_OFFSET,6),(write_line+TOTAL_DOWNTIME_ROW_OFFSET+1,6))

    chart = PieChart()

    chart.append(Series(values=yvalues,labels=xvalues))

    #chart.append(Series(labels,title="Total Uptime"))

    chart.drawing.height = 220
    chart.drawing.width = 350
    chart.drawing.top = (20*(write_line-1))+(i*5)
    sheet.add_chart(chart)

main()